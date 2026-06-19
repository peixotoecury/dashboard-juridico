"""
sync_prazos.py
Lê Prazos_e_agendas.xlsx, gera dados.json e faz upload para o Supabase.
Disparado automaticamente pelo watcher_prazos.py quando o Excel muda.
"""
import openpyxl, json, logging, requests
from datetime import datetime, date
from pathlib import Path

PASTA    = Path(__file__).parent
EXCEL    = Path(r"C:\Users\ach\OneDrive - Peixoto e Cury Advogados\Pastas - Time B\Controladoria\Base diária de atualização sócios\v_pc_controle_prazos_trabalhista.xlsx")
LOG      = PASTA / "sync_prazos.log"
JSON_OUT = PASTA / "dados.json"

# ── Supabase ──────────────────────────────────────────────────────────────────
SUPA_URL = "https://kgyeeuynerhpaqlhqwkx.supabase.co"
SUPA_KEY = "sb_publishable_MdUFFFJ5xRqiQysyxlWHFA_yk_LsknZ"
SUPA_TABELA = "prazos_agendas"

logging.basicConfig(filename=LOG, level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

def to_iso(val):
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    try:
        return str(val)[:10]
    except:
        return None

def ler_excel():
    import shutil, tempfile
    logging.info("Lendo Excel de prazos...")
    # Copia para temp para nao travar com Excel/OneDrive aberto
    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    shutil.copy2(str(EXCEL), str(tmp))
    wb = openpyxl.load_workbook(str(tmp), data_only=True, read_only=True)
    ws = wb.active
    headers = [str(c or "").strip() for c in
               next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    registros = []
    hoje = date.today()

    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))

        data_val = r.get("data")
        data_iso = to_iso(data_val)
        if not data_iso:
            continue

        # Só sincroniza registros dos últimos 60 dias até o futuro
        try:
            from datetime import timedelta
            data_dt2 = datetime.strptime(data_iso, "%Y-%m-%d").date()
            if data_dt2 < hoje - timedelta(days=60):
                continue
        except:
            pass

        # Pastas encerradas: inclui agendas sempre + prazos só se data >= hoje (Fatal/D-1/Futuro)
        # Prazos vencidos de encerrados são descartados
        if str(r.get("situacao_pasta") or "").strip().lower() == "encerrado":
            tipo_agenda = str(r.get("tipo_agenda") or "").strip()
            try:
                data_enc = datetime.strptime(data_iso, "%Y-%m-%d").date()
                eh_futuro = data_enc >= hoje  # Fatal (hoje) ou D-1/Futuro (futuro)
            except:
                eh_futuro = False
            if "Agenda" not in tipo_agenda and not eh_futuro:
                continue

        # Filtra apenas area TRABALHISTA
        area = str(r.get("area") or "").strip().upper()
        if area and "TRABALHISTA" not in area:
            continue

        # Descarta matriculas de sistemas automaticos (nao sao prazos humanos)
        MATRICULAS_DESCARTAR = {"PRAZO", "BLU", "TAC", "MSS", "AUDIENCIA", "CONTROL"}
        matricula = str(r.get("matricula") or "").strip().upper()
        if matricula in MATRICULAS_DESCARTAR:
            continue

        # Calcula dias úteis até o prazo (igual ao JavaScript do dashboard)
        try:
            data_dt = datetime.strptime(data_iso, "%Y-%m-%d").date()
            cal_diff = (data_dt - hoje).days
            if abs(cal_diff) > 14:
                dias_uteis = cal_diff
            elif cal_diff == 0:
                dias_uteis = 0
            else:
                step = 1 if cal_diff > 0 else -1
                count = 0
                cur = date(hoje.year, hoje.month, hoje.day)
                from datetime import timedelta
                while cur != data_dt:
                    cur += timedelta(days=step)
                    if cur.weekday() < 5:  # 0=seg ... 4=sex
                        count += step
                dias_uteis = count
            dias_restantes = dias_uteis
        except:
            dias_restantes = None

        # Status baseado em dias úteis — espelha exatamente o cálculo do dashboard
        if dias_restantes is None:
            status_calc = "Futuro"
        elif dias_restantes < -1:
            status_calc = "Vencido"
        elif dias_restantes == -1:
            status_calc = "Fatal"
        elif dias_restantes == 0:
            status_calc = "D-1"
        elif dias_restantes == 1:
            status_calc = "D-2"
        else:
            status_calc = "Futuro"

        registros.append({
            "data":             data_iso,
            "status":           status_calc,
            "situacao_pasta":   str(r.get("situacao_pasta") or "").strip(),
            "responsavel":      str(r.get("responsavel") or "").strip(),
            "grupo_processo":   str(r.get("grupo_processo") or "").strip(),
            "cliente":          str(r.get("cliente") or "").strip(),
            "tipo_agenda":      str(r.get("tipo_agenda") or "").strip(),
            "descricao_tarefa": str(r.get("descricao_tarefa") or "").strip(),
            "numero_processo":  str(r.get("numero_processo") or "").strip(),
            "parte_contraria":  str(r.get("parte_contraria") or "").strip(),
            "importante":       str(r.get("importante") or "").strip(),
            "matricula":        str(r.get("matricula") or "").strip(),
            "dias_restantes":   dias_restantes,
            # campos extras deste Excel
            "parte_principal":  str(r.get("parte_principal") or "").strip(),
            "area":             str(r.get("area") or "").strip(),
            "codigo_processo":  str(r.get("codigo_processo") or "").strip(),
            "codigo_agenda":    str(r.get("codigo_agenda") or "").strip(),
            "codigo_cliente":   str(r.get("codigo_cliente") or "").strip(),
            "quando_criado":    to_iso(r.get("quando_criado")),
            "quem_criou":       str(r.get("quem_criou") or "").strip(),
        })

    wb.close()
    try: tmp.unlink()
    except: pass
    logging.info(f"{len(registros)} prazos lidos.")
    return registros

def salvar_json(registros):
    payload = {
        "atualizado_em": datetime.now().isoformat(),
        "total": len(registros),
        "prazos": registros,
    }
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    logging.info(f"JSON salvo: {JSON_OUT}")

def upload_supabase(registros):
    headers = {
        "apikey": SUPA_KEY,
        "Authorization": f"Bearer {SUPA_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }

    TIMEOUT = 20  # segundos

    # Delete tudo usando id >= 0 (funciona com bigserial)
    r = requests.delete(
        f"{SUPA_URL}/rest/v1/{SUPA_TABELA}",
        headers={**headers, "Prefer": "return=minimal"},
        params={"id": "gte.0"},
        timeout=TIMEOUT,
    )
    logging.info(f"Delete Supabase: {r.status_code}")

    # Insere em lotes de 200
    LOTE = 200
    for i in range(0, len(registros), LOTE):
        lote = registros[i:i+LOTE]
        r = requests.post(
            f"{SUPA_URL}/rest/v1/{SUPA_TABELA}",
            headers=headers,
            json=lote,
            timeout=TIMEOUT,
        )
        if r.status_code in (200, 201):
            logging.info(f"Lote {i//LOTE+1}: {len(lote)} registros inseridos")
        else:
            logging.error(f"Erro Supabase lote {i//LOTE+1}: {r.status_code} {r.text[:200]}")

def main():
    logging.info("=== Início sync prazos ===")
    try:
        registros = ler_excel()
        salvar_json(registros)
        upload_supabase(registros)
        logging.info("=== Concluído ===")
        print(f"OK — {len(registros)} prazos sincronizados em {datetime.now().strftime('%H:%M:%S')}")
    except Exception as ex:
        logging.error(f"ERRO: {ex}", exc_info=True)
        print(f"ERRO: {ex}")

if __name__ == "__main__":
    main()
